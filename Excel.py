from openpyxl import load_workbook
from collections import namedtuple
import warnings


def get_analytes(workbook=None):
    wb = load_workbook(workbook)
    ws = wb.active
    this = [x[0] for x in ws.iter_rows(min_col=3, max_col=3, min_row=2, values_only=True)]
    return this


def getAreaValues(workbook):
    #Look into implementing a wrapper/ decorator here. Function body is too big. Look into a generator functional solution.
    wb = load_workbook(workbook)
    ws = wb.active
    try:
        headers = getIndices(ws)
    except StopIteration:
        raise ValueError
    name_idx, area_idx = getCols(ws, headers)
    if headers > 1:
        warnings.warn('Warning: Malformed columns', RuntimeWarning)
    if name_idx != 3 or area_idx != 10:
        warnings.warn('Warning: Malformed columns', RuntimeWarning)
    out = _getValues(ws, headers, name_idx, area_idx)
    j = 0
    for item in [x[1] for x in out]:
        if not isinstance(item, (int, float)):
            raise TypeError
        elif not item >= 0.0:
            raise ValueError
        elif item == 0.0:
            j += 1
        else:
            continue
    if j == len(out):
        warnings.warn('Warning: All values are zeroes', RuntimeWarning)
    return out


def _getValues(ws, headers, name_idx, area_idx):
    delta = area_idx - name_idx
    if delta >= 0:
        values = [[x[0], x[delta]] if x[delta] is not None or '' else [x[0], 0.0] for x in ws.iter_rows(
            min_row=headers+1, min_col=name_idx, max_col=area_idx, values_only=True)]
    else:
        values = [[x[delta], x[0]] if x[delta] is not None or '' else [x[0], 0.0] for x in ws.iter_rows(
            min_row=headers+1, min_col=area_idx, max_col=name_idx, values_only=True)]
    return values


def getIndices(ws):
    gen = ws.iter_rows(values_only=True)
    i = 1
    x = next(gen)
    while 'Area' not in x and 'Analyte' not in x:
        i = + 1
        x = next(gen)
    return i


def getCols(ws, i):
    x = [x for x in ws.iter_rows(min_row=i, max_row=i, values_only=True)]
    sentinel_dup_err = 0
    for idx, col in enumerate(x[0], start=1):
        if col == 'Area':
            sentinel_dup_err += 1
            area_idx = idx
        elif col == 'Analyte':
            name_idx = idx
    if sentinel_dup_err > 1:
        raise RuntimeError('Duplicate Areas columns')
    try:
        return (name_idx, area_idx)
    except UnboundLocalError:
        raise ValueError
